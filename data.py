from openpyxl import load_workbook, Workbook
from natsort import natsorted
import shutil
import glob
import re
import os
import cars_list as c


# TODO:ZROBIC KLASE NA PODSTWIE DOSTEPNYCH FUNKCJI

def create_file():
    wb = Workbook()
    ws = wb['Sheet']
    ws['A1'] = 'POJAZDY'
    ws['B1'] = 'ILOSC PALIWA'
    ws['C1'] = 'WAGA PALIWA'
    ws['D1'] = 'STAWKA'
    ws['E1'] = 'OBLICZENIA'
    ws['A21'] = 'PANDA,SKODA,SKODA_2'
    ws['A22'] = 'COURIER'
    ws['A23'] = 'LUBLIN'
    ws['A24'] = 'CITROEN'
    ws['A25'] = 'TRANSIT,TRANSIT_LEASING'
    ws['A26'] = 'UNIMOG,UNIMOG_2'
    ws['A27'] = 'FARMTRAC'
    ws['A28'] = 'NOREMAT,TYM,URSUS'
    ws['A29'] = 'REBAK,NAFRZEWNICA'
    ws['A30'] = 'SPRZET,REBAK_PB'

    # for row in range(1):
    #    ws1.append(range(0, 13))

    wb.save("calculation.xlsx")


def copy_files():
    source_dir = "/Users/btn/PycharmProjects/cars/2022/"
    dest_dir = "/Users/btn/PycharmProjects/environmental_protection/"
    files = glob.iglob(os.path.join(source_dir, "*.xlsx"))
    for file in files:
        if os.path.isfile(file):
            shutil.copy2(file, dest_dir)


def sort_cars():
    path = "/Users/btn/PycharmProjects/environmental_protection/"
    files = [f for f in glob.glob(path + "**/*.xlsx", recursive=True)]
    folders_show = re.compile(r".*(/environmental_protection/)")
    files_sort = natsorted([folders_show.sub('', x).strip() for x in files])
    print(files_sort)

    wb = load_workbook("calculation.xlsx")
    ws = wb['Sheet']
    ws.column_dimensions["A"].width = 25
    ws['A2'] = files_sort[0]
    ws['A3'] = files_sort[1]
    ws['A4'] = files_sort[2]
    ws['A5'] = files_sort[3]
    ws['A6'] = files_sort[4]
    ws['A7'] = files_sort[5]
    ws['A8'] = files_sort[6]
    ws['A9'] = files_sort[7]
    ws['A10'] = files_sort[8]
    ws['A11'] = files_sort[9]
    ws['A12'] = files_sort[10]
    ws['A13'] = files_sort[11]
    ws['A14'] = files_sort[12]
    ws['A15'] = files_sort[13]

    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 8
    wb.save("calculation.xlsx")


def cost():
    wb = load_workbook("calculation.xlsx")
    ws = wb['Sheet']

    ford = ws['A2'].value
    fc = float(input("Cost for " + ford + ": "))
    ws['D2'] = fc

    skoda = ws['A3'].value
    sf = float(input("Cost for " + skoda + ": "))
    ws['D3'] = sf

    skoda_2 = ws['A4'].value
    sf_2 = float(input("Cost for " + skoda_2 + ": "))
    ws['D4'] = sf_2

    fiat = ws['A5'].value
    fp = float(input("Cost for " + fiat + ": "))
    ws['D5'] = fp

    citroen = ws['A6'].value
    cj = float(input("Cost for " + citroen + ": "))
    ws['D6'] = cj

    daewoo = ws['A7'].value
    dl = float(input("Cost for " + daewoo + ": "))
    ws['D7'] = dl

    ford_2 = ws['A8'].value
    ft = float(input("Cost for " + ford_2 + ": "))
    ws['D8'] = ft

    ford_leasing = ws['A9'].value
    ft_2 = float(input("Cost for " + ford_leasing + ": "))
    ws['D9'] = ft_2

    farmtrac = ws['A10'].value
    f = float(input("Cost for " + farmtrac + ": "))
    ws['D10'] = f

    ursus = ws['A11'].value
    u = float(input("Cost for " + ursus + ": "))
    ws['D11'] = u

    tym = ws['A12'].value
    t = float(input("Cost for " + tym + ": "))
    ws['D12'] = t

    unimog = ws['A13'].value
    mu = float(input("Cost for " + unimog + ": "))
    ws['D13'] = mu

    unimog_2 = ws['A14'].value
    mu_2 = float(input("Cost for " + unimog_2 + ": "))
    ws['D14'] = mu_2

    noremat = ws['A15'].value
    n = float(input("Cost for " + noremat + ": "))
    ws['D15'] = n

    wb.save("calculation.xlsx")


def tabelka():
    wb = load_workbook("calculation.xlsx")
    ws = wb['Sheet']

    ws['B20'] = 'TABELKA D'
    ws['B21'] = '5'
    ws['B22'] = '6'
    ws['B23'] = '10'
    ws['B24'] = '11'
    ws['B25'] = '12'
    ws['B26'] = '18'
    ws['B27'] = '22'
    ws['B28'] = '23'
    ws['B29'] = '26'
    ws['B30'] = '27'

    wb.save("calculation.xlsx")


def calculations():
    wb = load_workbook("calculation.xlsx")
    ws = wb['Sheet']

    ws['E2'] = '=B2 * (C2 / 1000) * D2'
    ws['E3'] = '=B3 * (C3 / 1000) * D3'
    ws['E4'] = '=B4 * (C4 / 1000) * D4'
    ws['E5'] = '=B5 * (C5 / 1000) * D5'
    ws['E6'] = '=B6 * (C6 / 1000) * D6'
    ws['E7'] = '=B7 * (C7 / 1000) * D7'
    ws['E8'] = '=B8 * (C8 / 1000) * D8'
    ws['E9'] = '=B9 * (C9 / 1000) * D9'
    ws['E10'] = '=B10 * (C10 / 1000) * D10'
    ws['E11'] = '=B11 * (C11 / 1000) * D11'
    ws['E12'] = '=B12 * (C12 / 1000) * D12'
    ws['E13'] = '=B13 * (C13 / 1000) * D13'
    ws['E14'] = '=B14 * (C14 / 1000) * D14'
    ws['E15'] = '=B15 * (C15 / 1000) * D15'

    ws['E16'] = '=SUMA(E2:E15)'

    ws['C21'] = '=B2 * (C15 / 1000)'
    ws['C22'] = '=B3 * (C3 / 1000)'
    ws['C23'] = '=B4 * (C4 / 1000)'
    ws['C24'] = '=B5 * (C5 / 1000)'
    ws['C25'] = '=B6 * (C6 / 1000)'
    ws['C26'] = '=B7 * (C7 / 1000)'
    ws['C27'] = '=B8 * (C8 / 1000)'
    ws['C28'] = '=B9 * (C9 / 1000)'
    ws['C29'] = '=B10 * (C10 / 1000)'
    ws['C30'] = '=B11 * (C11 / 1000)'

    ws['D21'] = '=D3'
    ws['D22'] = '=D2'
    ws['D23'] = '=D7'
    ws['D24'] = '=D6'
    ws['D25'] = '=D8'
    ws['D26'] = '=D13'
    ws['D27'] = '=D10'
    ws['D28'] = '=D15'
    ws['D29'] = 'REBAK'
    ws['D30'] = 'SPRZET'

    ws['E21'] = '=C21*D21'
    ws['E22'] = '=C22*D22'
    ws['E23'] = '=C23*D23'
    ws['E24'] = '=C24*D24'
    ws['E25'] = '=C25*D25'
    ws['E26'] = '=C26*D26'
    ws['E27'] = '=C27*D27'
    ws['E28'] = '=C28*D28'
    # ws['E29'] = '=C29*D29'
    # ws['E30'] = '=C30*D30'

    ws['E31'] = '=SUMA(E21:E30)'

    wb.save("calculation.xlsx")


# copy_files()          #working
# create_file()         #working
# sort_cars()           #working
# cost()
tabelka()
calculations()

'''
c.ford()
c.skoda()
c.skoda_2()
c.fiat()
c.citroen()
c.daewoo()
c.ford_2()
c.ford_leasing()
c.farmtrac()
c.ursus()
c.tym()
c.unimog()
c.unimog_2()
c.noremat()
'''

'''
def copy_datas():
    # z listy skopiowac nazwy pojazdow do komorek
    wb = load_workbook("1.FORD COURIER.xlsx")
    sheet = wb["ochrona_srodowiska"]

    wb1 = load_workbook("calculation.xlsx")
    sheet1 = wb1["Sheet1"]

    # Copy range of cells as a nested list
    # Takes: start cell, end cell, and sheet you want to copy from.
    def copyRange(startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        # Loops through selected Rows
        for i in range(startRow, endRow + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol, endCol + 1, 1):
                rowSelected.append(sheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        return rangeSelected

    # Paste range
    # Paste data from copyRange into template sheet
    def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
        countRow = 0
        for i in range(startRow, endRow + 1, 1):
            countCol = 0
            for j in range(startCol, endCol + 1, 1):
                sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1

    def createData():
        print("Processing...")
        selectedRange = copyRange(2, 13, 2, 13, sheet)  # Change the 4 number values
        pastingRange = pasteRange(1, 3, 4, 15, sheet1, selectedRange)  # Change the 4 number values
        # You can save the template as another file to create a new file here too.s
        wb1.save("calculation.xlsx")
'''
