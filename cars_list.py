from openpyxl import load_workbook


def ford():
    wb = load_workbook("1.FORD COURIER.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B2'] = cell.value
    ws1['C2'] = float(0.8333)

    wb1.save("calculation.xlsx")


def skoda():
    wb = load_workbook("2.SKODA FABIA.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B3'] = cell.value
    ws1['C3'] = float(0.7475)

    wb1.save("calculation.xlsx")


def skoda_2():
    wb = load_workbook("3.SKODA FABIA 2.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B4'] = cell.value
    ws1['C4'] = float(0.7475)

    wb1.save("calculation.xlsx")


def fiat():
    wb = load_workbook("4.FIAT PANDA.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B5'] = cell.value
    ws1['C5'] = float(0.7475)

    wb1.save("calculation.xlsx")


def citroen():
    wb = load_workbook("5.CITROEN JUMPER.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B6'] = cell.value
    ws1['C6'] = float(0.8333)

    wb1.save("calculation.xlsx")


def daewoo():
    wb = load_workbook("6.DAEWOO LUBLIN.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B7'] = cell.value
    ws1['C7'] = float(0.8333)

    wb1.save("calculation.xlsx")


def ford_2():
    wb = load_workbook("7.FORD TRANSIT.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B8'] = cell.value
    ws1['C8'] = float(0.8333)

    wb1.save("calculation.xlsx")


def ford_leasing():
    wb = load_workbook("8.FORD TRANSIT LEASING.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']
    cell = ws['B13']
    ws1['B9'] = cell.value
    ws1['C9'] = float(0.8333)

    wb1.save("calculation.xlsx")


def farmtrac():
    wb = load_workbook("9.FARMTRAC.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B10'] = cell.value
    ws1['C10'] = float(0.8333)

    wb1.save("calculation.xlsx")


def ursus():
    wb = load_workbook("10.URSUS.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B11'] = cell.value
    ws1['C11'] = float(0.8333)

    wb1.save("calculation.xlsx")


def tym():
    wb = load_workbook("11.TYM.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B12'] = cell.value
    ws1['C12'] = float(0.8333)

    wb1.save("calculation.xlsx")


def unimog():
    wb = load_workbook("12.UNIMOG.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B13'] = cell.value
    ws1['C13'] = float(0.8333)

    wb1.save("calculation.xlsx")


def unimog_2():
    wb = load_workbook("13.UNIMOG 2.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B14'] = cell.value
    ws1['C14'] = float(0.8333)

    wb1.save("calculation.xlsx")


def noremat():
    wb = load_workbook("14.NOREMAT.xlsx", data_only=True)
    ws = wb['ochrona_srodowiska']
    wb1 = load_workbook("calculation.xlsx")
    ws1 = wb1['Sheet']

    cell = ws['B13']
    ws1['B15'] = cell.value
    ws1['C15'] = float(0.8333)

    wb1.save("calculation.xlsx")